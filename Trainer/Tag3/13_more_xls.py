from openpyxl import Workbook, load_workbook

def create_xls(filepath):
    wb = Workbook()
    wb.save(filepath)

def write_xls(filepath, dictionary):
    wb = load_workbook(filepath)
    sheet = wb.active

    headers = [x for x in dictionary[0]]

    for index, value in enumerate(headers):
        sheet.cell(row=1, column=index+1).value = value

    for i, x in enumerate(dictionary):
        for idx,value in enumerate(x.values()):
            sheet.cell(row=i+2, column=idx+1).value = value

    wb.save(filepath)